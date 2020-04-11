from openpyxl import load_workbook
from common import read_path
from common.read_config import ReadConfig
from common.do_mysql import DoMySql


class DoExcel:
    def __init__(self, file_name):
        self.file_name = file_name
        self.wb = load_workbook(self.file_name)

    def get_title(self, sheet_name):
        """获取每一列数据的标题"""
        sheet = self.wb[sheet_name]
        title = []
        for i in range(1, sheet.max_column + 1):
            title.append(sheet.cell(1, i).value)
        return title

    def get_test_data(self, sheet_name):
        """获取测试数据"""
        sheet = self.wb[sheet_name]
        test_data = []
        title = self.get_title(sheet_name)
        # 获取最大手机号
        sql_max_tel = 'select max(mobilephone) from member'
        tel = int(DoMySql().do_my_sql(sql_max_tel, 1)[0]) + 1
        # 存储已使用的手机号
        self.used_tel(tel)
        self.used_tel(tel + 1)
        # 测试运行的模式
        mode = ReadConfig().read_config(read_path.conf_path, 'MODE', 'mode')
        # 测试运行的用例编号列表
        case_id_list = ReadConfig().read_config(read_path.conf_path, 'MODE', 'case_id_list')
        # 按行读取，从第2行开始
        for i in range(2, sheet.max_row + 1):
            row_data = {}
            # 按列读取，从第1列开始
            for j in range(1, sheet.max_column + 1):
                row_data[title[j - 1]] = sheet.cell(i, j).value

            # 变量的替换
            if row_data['Param'].find('${tel}') != -1:
                row_data['Param'] = row_data['Param'].replace('${tel}', str(tel))
            elif row_data['Param'].find('${tel+1}') != -1:
                row_data['Param'] = row_data['Param'].replace('${tel+1}', str(tel + 1))
            elif row_data['Param'].find('${tel+2}') != -1:
                row_data['Param'] = row_data['Param'].replace('${tel+2}', str(tel + 2))
            test_data.append(row_data)
        # 如果mode=1，执行所有测试用例
        if mode == '1':
            final_data = test_data
        # 如果mode !=1，执行case_id_list里面的用例
        else:
            final_data = []
            for item in test_data:
                if item['Case_Id'] in eval(case_id_list):
                    final_data.append(item)
        return final_data

    def write_back(self, row, col, new_value):
        """将测试结果写回测试用例文件"""
        sheet = self.wb['test_data']
        sheet.cell(row, col).value = new_value
        self.wb.save(self.file_name)

    def used_tel(self, used_tel):
        """将已使用的手机号存到文件中"""
        sheet = self.wb['used_tel']
        sheet.cell(sheet.max_row + 1, 1).value = used_tel
        self.wb.save(self.file_name)


if __name__ == '__main__':
    data = DoExcel(read_path.test_data_path).get_test_data('test_data')
    print(data)